import os
from flask import Flask, flash, session, current_app, send_from_directory
from flask import render_template, redirect, url_for, request, get_flashed_messages
from flask_bcrypt import generate_password_hash, check_password_hash
from peewee import *
import pandas as pd
from openpyxl import *
import random
from string import ascii_lowercase
from werkzeug.utils import secure_filename

from profil import db, Guest, Player, Orga, Tables
from Guest_route import bp as guest_bp
from Orga_route import bp as orga_bp
from Orga_route import orga_login_required_id
from Guest_route import login_required_id
from create_wed import CreateMariage
from send_email import *
from form import ExchangeSeats, ChangePassword, nb_guest_couple, TablesFakeMariage, ResetPassword, ContactUs, UploadFile
import RO

from urllib.parse import quote, unquote

def Chemin(folder_name):

    return os.path.join(os.path.abspath(os.path.dirname(folder_name)), folder_name)


def create_app(static_folder_name, templates_folder_name):
    app = Flask(__name__)  #creating the app
    app.secret_key = 'superbe_secret_key'
    app.static_folder = Chemin(static_folder_name)  #giving flask the path to the static files
    app.template_folder = Chemin(templates_folder_name)  #giving flask the path to the templates
    app.register_blueprint(guest_bp)  # Blueprint of the app for the guests' routes
    app.register_blueprint(orga_bp)  # Blueprint of the app for the organizers' routes
    db.connect()  #connecting to the database
    db.create_tables([Guest, Player, Orga, Tables])  #creating the tables 
    return app

app = create_app("static", "templates")


def get_orga_player_email(statut, email, mariage):
    if statut == 'guest': 
        try: 
            guest = Player.get((Player.mariage==mariage) & (Player.email == email))
        except DoesNotExist: 
            return None
        return guest
    if statut == 'orga': 
        try: 
            orga = Orga.get((Orga.mariage==mariage) & (Orga.email == email))
        except DoesNotExist: 
            return None
        return orga
    else:
        return None


def get_orga_guest(statut, ident): 
    if statut == 'guest': 
        try: 
            guest = Player.get(Player.id == ident)
        except DoesNotExist: 
            return ("L'invité qui essaye de se connecter n'est pas dans la base de données")
        return guest
    if statut == 'orga': 
        try: 
            orga = Orga.get(Orga.id == ident)
        except DoesNotExist: 
            return ("L'invité qui essaye de se connecter n'est pas dans la base de données")
        return orga
    else:
        return None


def get_orga_guest_2(statut, ident):
    if statut == 'guest':
        try: 
            player = Player.get(Player.id == ident)
        except DoesNotExist:
            flash("L'utilisateur qui essaye de ne connecter n'est pas dans la base de données")
            return redirect(url_for('logout', statut=statut, ident=ident))
        try: 
            orga = Orga.get(Orga.mariage == player.mariage)
        except DoesNotExist:
            flash("Aucun utilisateur n'est organisateur de ce mariage")
            return redirect(url_for('logout', statut=statut, ident=ident))
        return player, orga
    if statut == 'orga':
        try: 
            orga = Orga.get(Orga.id == ident)
        except DoesNotExist:
            flash("L'utilisateur qui essaye de ne connecter n'est pas dans la base de données")
            return redirect(url_for('logout', statut=statut, ident=ident))
        return orga
    else: 
        return None
    

@app.route('/logout/<statut>/<ident>')
def logout(statut, ident):
    # Delete the user session
    session.pop(str(statut + str(ident)), None)
    print(current_app.root_path)
    print(os.path.join(os.path.abspath(os.path.dirname('src')), 'src'))
    # Redirect the user to the accueil page
    return redirect(url_for('page_de_garde'))
    

@app.route('/')
def page_de_garde():
    """
    This function 'page_de_garde' handle the back-end of accueil page.
    :return: Page d'accueil
    :rtype: template
    """
    return render_template('accueil.html', title = 'Accueil', log=False, messages = get_flashed_messages() )


@app.route('/<statut>/<ident>')
def page_de_garde_login(statut, ident):
    @login_required_id(ident)
    def wrapper():
        if statut == 'guest':
            player, orga = get_orga_guest_2(statut, ident)
            return render_template('accueil.html', title='Accueil', log = True, ident=ident, mariage=player.mariage, full_name = player.full_name(),
                                   statut=statut, groom_full_name = orga.groom_full_name(), bride_full_name = orga.bride_full_name())
        if statut == 'orga':
            orga = get_orga_guest_2('orga', ident)
            return render_template('accueil.html', title='Accueil',log = True, ident=ident, mariage=orga.mariage, full_name = orga.orga_full_name(),
                                   statut='orga', groom_full_name = orga.groom_full_name(), bride_full_name=orga.bride_full_name())
        else: 
            return "non"
    return wrapper()


@app.route('/about_us')
def about_us():
    """
    This function 'about_us' handle the back-end of about us page.

    :return: About us page 
    :rtype: template
    """
    return render_template('about_us.html', title = "about us", log=False)


@app.route('/about_us/<statut>/<ident>')
def about_us_login(statut, ident):
    @login_required_id(ident)
    def wrapper():
        if statut == 'guest':
            player, orga = get_orga_guest_2(statut, ident)
            return render_template('about_us.html', title='About us', log = True, ident=ident, mariage=player.mariage, full_name = player.full_name(), 
                                   statut=statut, groom_full_name = orga.groom_full_name(), bride_full_name = orga.bride_full_name())
        else:
            orga = get_orga_guest_2(statut, ident)
            return render_template('about_us.html', title='About us',log = True, ident=ident, mariage=orga.mariage, full_name = orga.orga_full_name(), 
                                   statut='orga', groom_full_name = orga.groom_full_name(), bride_full_name=orga.bride_full_name())
    return wrapper()


@app.route('/algo')
def algo():
    """
    This function 'algo' handle the back-end of "how does it works ?" page.

    :return: Algo page 
    :rtype: template
    """
    return render_template('algo.html', title = "Algo", log=False)


@app.route('/algo/<statut>/<ident>')
def algo_login(statut, ident):
    @login_required_id(ident)
    def wrapper():
        if statut == 'guest':
            player, orga = get_orga_guest_2(statut, ident)
            return render_template('algo.html', title='Algo', log = True, ident=ident, mariage=player.mariage, full_name = player.full_name(),
                                    statut=statut, groom_full_name = orga.groom_full_name(), bride_full_name = orga.bride_full_name())
        else:
            orga = get_orga_guest_2(statut, ident)
            return render_template('algo.html', title='Algo',log = True, ident=ident, mariage=orga.mariage, full_name = orga.orga_full_name(),
                                    statut='orga', groom_full_name = orga.groom_full_name(), bride_full_name=orga.bride_full_name())
    return wrapper()


@app.route('/contact_us', methods=['GET','POST'])
def contact_us():
    """
    This function 'contact_us' handle the back-end of "contact_us" page.

    :return: contact_us page 
    :rtype: template
    """
    #In odrer to modify the list of people that receive mail from people contacting us you need to modify communication list
    communication_list=['pierrelouis.legoff@free.fr']
    message=''
    form = ContactUs()
    if request.method == 'POST':
        if form.validate_on_submit():            
            message = str('L email a été envoyé par '+str(form.name.data)+ ' Si vous voulez le contacter voila son adresse mail ' + str(form.email.data) +
                          '\n Voici son message : \n' + str(form.body.data))
            send_mail_organisation(message,form.subject.data,communication_list)
            return redirect(url_for('page_de_garde'))
    return render_template('contact_us.html', title = "Nous contacter", log=False, form=form)


@app.route('/contact_us/<statut>/<ident>', methods=['GET', 'POST'])
def contact_us_login(statut, ident):
    @login_required_id(ident)
    def wrapper():
        communication_list=['pierrelouis.legoff@free.fr']
        message=''
        form = ContactUs()
        if statut == 'guest':
            player, orga = get_orga_guest_2(statut, ident) 
            if request.method == 'POST':
                if form.validate_on_submit():           
                    message = str('L email a été envoyé par '+str(form.name.data)+ ' Si vous voulez le contacter voila son adresse mail ' + str(form.email.data) +'\n Voici son message : \n' + str(form.body.data))
                    send_mail_organisation(message,form.subject.data,communication_list)
                    return redirect(url_for('page_de_garde_login', statut=statut, ident=ident))
            return render_template('contact_us.html', title='Nous contacter', log = True, ident=ident, form=form, mariage=player.mariage, full_name = player.full_name(), 
                                   statut=statut, groom_full_name = orga.groom_full_name(), bride_full_name = orga.bride_full_name())
        else:
            orga = get_orga_guest_2(statut, ident)
            if request.method == 'POST':
                if form.validate_on_submit():           
                    message = str('L email a été envoyé par '+str(form.name.data)+ ' Si vous voulez le contacter voila son adresse mail ' + str(form.email.data) +'\n Voici son message : \n' + str(form.body.data))
                    send_mail_organisation(message,form.subject.data,communication_list)
                    return redirect(url_for('page_de_garde_login', statut=statut, ident=ident))
            return render_template('contact_us.html', title='Nous contacter',log = True, ident=ident, form=form, mariage=orga.mariage, full_name = orga.orga_full_name(), 
                                   statut='orga', groom_full_name = orga.groom_full_name(), bride_full_name=orga.bride_full_name())
    return wrapper()


@app.route('/change_password/<statut>/<mariage>/<ident>', methods=['get', 'post'])
def change_password(statut, mariage, ident):
    @login_required_id(ident)
    def wrapper():
        form = ChangePassword()
        if request.method == 'POST':
            if form.validate_on_submit():
                user = get_orga_guest(statut, ident)
                if check_password_hash(user.password, form.actual_password.data):
                    if form.must_be_equal():
                        user.password = generate_password_hash(form.new_password.data).decode('utf-8')
                        user.save()
                        if statut == 'guest':
                            flash("Votre mot de passe a bien été modifié")
                            return redirect(url_for('guest.guest_profil', statut = statut, ident = ident, mariage = mariage, full_name=user.full_name(), guest_ident = ident))
                        if statut == 'orga':
                            return redirect(url_for('orga.orga_info', statut=statut, mariage=mariage, ident=ident))
                        else:
                            flash("Erreur, veuillez vous reconecté")
                            return redirect(url_for('logout', statut=statut, ident=ident))
                    else:
                        flash("Les deux mots de passe doivent être les mêmes")
                        return redirect(url_for('change_password', statut=statut, mariage=mariage, ident=ident))
                else: 
                    flash("mot de passe incorrect")
                    return redirect(url_for('change_password', statut=statut, mariage=mariage, ident=ident))
            else:
                flash("Les champs sont mal renseignés, veuillez rééssayer")
                return redirect(url_for('change_password', statut=statut, mariage=mariage, ident=ident))
        return render_template('change_password.html', form=form, statut=statut, ident=ident, messages=get_flashed_messages())
    return wrapper()


@app.route('/password_forgotten/<status>', methods=['get','post'])
def password_forgotten(status):
    message=''
    form = ResetPassword()
    if request.method == 'POST':
        if form.validate_on_submit():
            subject='Mariage Seating&CO : Réinitialisation de votre mot de passe'
            mariage=form.mariage.data
            email=form.email.data
            communication_list=[email]
            print(email)
            print(communication_list)
            user = get_orga_player_email(status, email, mariage)
            if user is None:
                flash("Vous n'avez pas de compte, vous ne pouvez pas changer votre mot de passe. Vérifier les informations que vous avez renseigné")
                return redirect(url_for('page_de_garde'))
            else:
                code_reset = ''.join(random.choice(ascii_lowercase) for i in range(8))
                print(code_reset)
                session[email] = code_reset
                mariage_endoded = quote(mariage)
                email_encoded = quote(email)
                status_encoded = quote(status)
                message= str('Votre code pour réinitialiser votre mot de passe est le suivant: ' + str(code_reset) + "\n Vous devez le renseigner en cliquant sur le lien suivant : http://127.0.0.1:5000/reset_password/" + mariage_endoded + '/' + email_encoded + '/' + status_encoded)
                send_mail_organisation(message,subject,communication_list)
                return redirect(url_for('reset_password', mariage = mariage_endoded, email=email_encoded, status=status_encoded))
    return render_template('password_forgotten.html', form=form, messages=get_flashed_messages())


@app.route('/reset_password/<mariage>/<email>/<status>', methods=['get','post'])
def reset_password(mariage, email, status):
    mariage_decoded = unquote(mariage)
    email_decoded = unquote(email)
    status_decoded = unquote(status)
    form = ChangePassword()
    if request.method == 'POST':
        if form.validate_on_submit():
            if form.verify_code(session[email_decoded]):
                if form.must_be_equal():
                    user = get_orga_player_email(status_decoded, email_decoded, mariage_decoded)
                    user.password = generate_password_hash(form.new_password.data).decode('utf-8')
                    user.save()
                    session.pop(email_decoded, None)
                    if status_decoded == 'guest':
                        flash("Votre mot de passe a bien été modifié")
                        return redirect(url_for('guest.log_guest'))
                    if status_decoded == 'orga':
                        flash("Votre mot de passe a bien été modifié")
                        return redirect(url_for('orga.log_orga'))
                else: 
                    flash("Les deux mots de passe doivent être les mêmes")
                    return redirect(url_for('reset_password', mariage=mariage, status=status, email=email))
            else: 
                flash("Le code renseigné est mauvais")
                return redirect(url_for('reset_password', mariage=mariage, status=status, email=email))
        else: 
            flash("Les informations sont mal renseignées")
            return redirect(url_for('reset_password', mariage=mariage, status=status, email=email))

    return render_template('reset_password.html', form=form, messages=get_flashed_messages())


@app.route('/add_guest/<mariage>/<statut>/<ident>', methods=['get','post'])
def add_guest(mariage, statut, ident):
    """
    This function 'add_guest' handle the back-end of the add_guest page : it allows an organizers to give its guest's list.

    :param mariage: name of the mariage of the guest accessing to its personnal profil
    :type mariage: str
    
    :return: Page of the add_guest 
    :rtype: template
    """ 
    @orga_login_required_id(ident)
    def wrapper():
        form = UploadFile()
        if request.method == 'POST':
            if form.validate_on_submit():
                #first, we create a folder for this wedding 
                dossier_path = os.path.join(app.static_folder, mariage)
                if not os.path.isdir(dossier_path):
                    os.mkdir(dossier_path)
                
                #file with the guest list
                file_guest = form.file_guest.data 
                file_guest.save(os.path.join(dossier_path, secure_filename(file_guest.filename)))
                file_guest_path = os.path.join(dossier_path, file_guest.filename)
                #checking is the file is an excel
                if not file_guest_path.endswith('.xlsx'):
                    flash("Vous ne pouvez importer que des fichiers excel")
                    return redirect(url_for('add_guest', statut=statut, ident=ident, mariage=mariage))
                
                # file with the table list
                file_tables = form.file_tables.data 
                file_tables.save(os.path.join(dossier_path, secure_filename(file_tables.filename)))
                file_tables_path = os.path.join(dossier_path, file_tables.filename)
                if not file_tables_path.endswith('.xlsx'):
                    flash("Vous ne pouvez importer que des fichiers excel")
                    return redirect(url_for('add_guest', statut=statut, ident=ident, mariage=mariage))
                
                #trying to import the files
                try:   
                    wb_guest = load_workbook(filename= file_guest_path , read_only=True)
                    wb_tables = load_workbook(filename=file_tables_path, read_only=True)
                except FileNotFoundError:
                    flash("Le fichier spécifié est introuvable.")
                    return redirect(url_for('add_guest', mariage=mariage, statut=statut, ident=ident))
                except:
                    flash("Une erreur s'est produite lors de l'enregistrement du fichier. Vérifier que le fichier est au bon format.")
                    return redirect(url_for('add_guest', mariage=mariage, statut=statut, ident=ident))
            
                #treating the guests' file
                sheet_guest = wb_guest.active
                rows_g = sheet_guest.max_row 
                cols_g = sheet_guest.max_column
                first_row_g = 1
                for j in range (1, rows_g):
                    if sheet_guest.cell(row = j, column = cols_g-2).value is not None: 
                        first_row_g = j
                        break

                for j in range (first_row_g +1, rows_g+1):
                    guest_email = sheet_guest.cell(row = j, column = cols_g).value 
                    if Guest.get_guest_mail(mariage, guest_email) is None:
                        guest = Guest.create(
                            name = sheet_guest.cell(row = j, column = cols_g-2 ).value,
                            surname = sheet_guest.cell(row = j, column = cols_g -1).value, 
                            email = guest_email,
                            mariage = mariage
                        )
                        guest.save()
                # treating the tables' file 
                sheet_tables = wb_tables.active
                rows_t = sheet_tables.max_row 
                cols_t = sheet_tables.max_column
                first_row_t = 1
                for j in range (1, rows_t):
                    if sheet_tables.cell(row = j, column = cols_t-1).value is not None: 
                        first_row_t = j
                        break

                for j in range(first_row_t + 1, rows_t+1):
                    table_capacity = sheet_tables.cell(row = j, column = cols_t-1 ).value
                    table_number = sheet_tables.cell(row = j, column = cols_t ).value
                    table = Tables.get_table_capa(mariage, table_capacity) 
                    if table is None:    
                        new_table = Tables.create(
                            mariage = mariage,
                            capacity = table_capacity,
                            number = table_number
                        )
                        new_table.save()
                    else: 
                        table.number = table_number
                        table.save()

                flash("Le fichier a bien été traité et enregistré. Vous pouvez désormais aller consulter votre liste d'inviter dans l'onglet correspondant !")
                return redirect(url_for('add_guest', mariage=mariage, statut=statut, ident=ident))  
            else:
                flash("Un des champs renseigné n'est pas valide")
                return redirect(url_for('add_guest', mariage=mariage, statut=statut, ident=ident))
        else:
            return render_template('orga_add_guest.html', form=form, statut=statut, ident=ident, mariage=mariage, messages=get_flashed_messages())
    return wrapper()


@app.route('/tables_assignment/<statut>/<mariage>/<ident>', methods=['GET', 'POST'])
def table_assigment(statut,mariage, ident):
    """
    This function 'table_assignment' handle the back-end of the table assignment page :
    it allows an organizers to generate a table assignement for the players of his wedding.

    :param mariage: name of the mariage of the guest accessing to its personnal profil
    :type mariage: str
    
    :return: Page of tables table assignment 
    :rtype: template
    """
    @orga_login_required_id(ident)
    def wrapper():
        form = ExchangeSeats(mariage)
        if request.method == 'POST':
            if form.validate_on_submit():
                id_1 = form.player_1.data
                id_2 = form.player_2.data
                if id_1 == 0 or id_2 == 0:
                    flash("Veuillez sélectionner 2 invités pour effectuer l'échange")
                    return redirect(url_for('table_assigment', mariage=mariage, statut=statut, ident=ident))
                else:
                    player_1 = Player.get_player_id(id_1)
                    player_2 = Player.get_player_id(id_2)
                    if player_1 == None or player_2 == None: 
                        flash("L'une des deux personnes sélectrionnée n'est pas invité à ce mariage.")
                        return redirect(url_for('table_assigment', mariage=mariage, statut=statut, ident=ident))
                    
                    table_1 = player_1.table
                    table_2 = player_2.table
                    player_1.table = table_2
                    player_1.save()
                    player_2.table = table_1
                    player_2.save()
                    return redirect(url_for('orga.seating_plan', statut=statut, mariage=mariage, ident=ident))

        # We have to make sure there is no table with 0 capacity
        for table in Tables.select().where(Tables.mariage == mariage):
            if table.capacity == 0:
                table.delete_instance()

        event = RO.Event(mariage)
        P = RO.PLNE_solver(event, mariage)
        PLNE = P.seating_plan()
        feasibility, repartition = PLNE[0], PLNE[1]
        if feasibility == 0:
            flash(str("Impossible de générer un plan de table avec " + str(repartition[0]) + ' invités, et ' + str(repartition[1]) + ' places.' "\n Veuillez ajouter des places pour générer un plan de table."))
            return redirect(url_for('orga.guest_list', statut=statut,mariage=mariage, ident=ident))
        nb_table = len(repartition)
        table_assignement = []
        nb_people_side_fav = [0, 0, 0]
        query = Player.select().where((Player.mariage == mariage) & (Player.table.is_null(False)))
        for player in query:
            table_assignement.append([player.color(),player.table])
            if player.same_table(player.favorite_guest_1):
                nb_people_side_fav[0] += 1
            if player.same_table(player.favorite_guest_2):
                nb_people_side_fav[1] += 1
            if player.same_table(player.favorite_guest_3):
                nb_people_side_fav[2] += 1
        nb_guest = len(table_assignement)

        nb_couple_side_by_side = 0
        couple = Player.select().where((Player.mariage == mariage) & (Player.couple_situation == 1))
        nb_couple = couple.count()/2
        for player in couple:
            if player.same_table(player.partner_name):
                nb_couple_side_by_side += 1
        return render_template("orga_table_assigment.html", statut=statut, ident=ident, mariage=mariage, seating_plan_bool=True, table_assignement=table_assignement,
                               nb_table=nb_table, nb_guest=nb_guest, form=form, nb_couple_side_by_side= int(nb_couple_side_by_side/2), nb_couple=int(nb_couple), nb_people_side_fav=nb_people_side_fav, messages=get_flashed_messages() )
    return wrapper()


@app.route('/uploads/<path:filename>/<ident>/<mariage>/<statut>', methods=['GET', 'POST'])
def download(filename, ident, mariage, statut):
    # Appending app path to upload folder path within app root folder
    uploads = os.path.join(current_app.root_path, app.static_folder, 'upload')
    # Returning file from appended path
    return send_from_directory(directory=uploads, filename=filename)


@app.route('/upload_seating_plan/<ident>/<mariage>/<statut>', methods=['GET', 'POST'])
def download_seating_plan(ident, mariage, statut):
    # Appending app path to upload folder path within app root folder
    dossier_path = os.path.join(app.static_folder, mariage, "upload_seating_plan")
    excel_file = str('seating_plan_' + mariage + '.xlsx')
    path_file = os.path.join(dossier_path, excel_file)
    nb_table = 0
    for table in Tables.select().where(Tables.mariage == mariage):
        nb_table += table.number
    dict_table = {}
    ### attention ValueError: All arrays must be of the same length !! 
    ## on récupère alors le maximum des capacite, on créer des liste de cette taille la et on ne rempli
    ## que les cases concernées 
    max_value = Tables.select(fn.MAX(Tables.capacity)).where(Tables.mariage == mariage).scalar()
    for i in range (nb_table):
        table_i = [None]*max_value
        j = 0
        for player in Player.select().where((Player.mariage ==mariage) & (Player.table == i)):
            table_i[j] = player.full_name()
            j+=1
        dict_table["table " + str(i)] = table_i
    if os.path.isdir(dossier_path):
        os.remove(path_file)
    else:
        os.mkdir(dossier_path)
    df = pd.DataFrame(dict_table)
    df.to_excel(path_file, index=False)
    return send_from_directory(directory=dossier_path, filename=excel_file)


@app.route('/create_fake_mariage/', methods=['GET', 'POST'])
def create_fake_mariage():
    form = nb_guest_couple()
    if request.method == 'POST':
        if form.validate_on_submit():
            if form.name_orga_available():
                if form.name_player_available():
                    if form.verify_nb_couple():
                                return redirect(url_for("generate_wedding_nb_tables", mariage_name=form.name.data, nb_guest=form.nb_guest.data,
                                                         nb_couple=form.nb_couple.data))
                    else:
                        flash("Le nombre de couples rensigné est trop grand par rapport au nombre d'invités \n Veuillez renseigner un nombre de couples inférieur au double du nombre d'invités")
                        return redirect(url_for("create_fake_mariage"))
            else: 
                flash("Le nom de mariage que vous avez renseigné n'est pas disponible, veuillez en choisir un autre")
                return redirect(url_for("create_fake_mariage"))
    return render_template("generate_wedding.html", form=form,is_generated=False, messages=get_flashed_messages())


@app.route('/create_fake_mariage_nb_tables/<mariage_name>/<nb_guest>/<nb_couple>', methods=['GET', 'POST'])
def generate_wedding_nb_tables(mariage_name, nb_guest, nb_couple):
    form = TablesFakeMariage(nb_guest)
    if request.method == 'POST':
        if form.validate_on_submit():
            # careful : the form.field.data are string, we have to recreate lists to utilize them
            list = form.table_capacity_number.data.strip("[]").split("], [")
            tables = [[int(x) for x in sub_list.split(", ")] for sub_list in list]
            print(tables)
            mar = CreateMariage(int(nb_guest), tables, mariage_name, int(nb_couple))
            mar.create_player()
            mar.create_tables()
            email_player = mar.guest_not_in_couple_isnt_coming() # deleting one player from the database, not in a couple for simplicity
            return redirect(url_for('fake_mariage_created', email_player=email_player, mariage_name=mariage_name, nb_couple=nb_couple, nb_guest=nb_guest))
    return render_template("generate_wedding_nb_tables.html", form=form, mariage_name=mariage_name, nb_guest=nb_guest)


@app.route('/fake_mariage_created/<email_player>/<mariage_name>/<nb_guest>/<nb_couple>', methods=['GET', 'POST'])
def fake_mariage_created(email_player, mariage_name, nb_guest, nb_couple):
    form = nb_guest_couple()
    if request.method == 'POST':
        if form.validate_on_submit():
            if form.name_orga_available():
                if form.name_player_available():
                    if form.verify_nb_couple():
                                return redirect(url_for("generate_wedding_nb_tables", mariage_name=form.name.data, nb_guest=form.nb_guest.data, nb_couple=form.nb_couple.data))
                    else:
                        flash("Le nombre de couple rensigné est trop grand par rapport aux nombre d'invités \n Veuillez renseigner un nombre de couple inférieur au double du nombre d'invités")
                        redirect(url_for('fake_mariage_created', email_player=email_player, mariage_name=mariage_name, nb_couple=nb_couple, nb_guest=nb_guest))
            else: 
                flash("Le nom de mariage que vous avez renseigné n'est pas disponible, veuillez en choisir un autre")
                redirect(url_for('fake_mariage_created', email_player=email_player, mariage_name=mariage_name, nb_couple=nb_couple, nb_guest=nb_guest))
    return render_template("generate_wedding.html", form=form,is_generated=True, email_player=email_player,
                            mariage_name=mariage_name, nb_couple=nb_couple, nb_guest=nb_guest, messages=get_flashed_messages())


db.close()

if __name__ == "__main__":
    app.run(debug=True)
